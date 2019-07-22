#Below code is for fetching the email ids of those customers who have not returned their tools on their respective return dates.
import numpy as np
import pandas as pd

df=pd.read_excel('dummy.xlsx')

from datetime import date
today = date.today()

to=list(df[df['Date']<today]['Email'])
to

#Below code is for sending mails as a reminder.
# Python code to illustrate Sending mail  
# to multiple users  
# from your Gmail account  
import smtplib 
  
for i in range(len(to)): 
    s = smtplib.SMTP('smtp-mail.outlook.com', 587) 
    s.starttls() 
    s.login("piyushchauhan0311@outlook.com", "piyush@03") 
    message = """
Subject: Remainder regarding the return of machines.

This is a test e-mail message.
""" 
    #message = "Message_you_need_to_send"
    s.sendmail("piyushchauhan0311@outlook.com", to[i], message) 
    s.quit()

#Below code is for fetching the comment by customer.
import imaplib
import email

# Connect to imap server
username = 'piyushchauhan0311@outlook.com'
password = 'piyush@03'
mail = imaplib.IMAP4_SSL('outlook.office365.com')
mail.login(username, password)

# retrieve a list of the mailboxes and select one
result, mailboxes = mail.list()
mail.select("inbox")

import openpyxl

for j in range(len(to)):
    type, data = mail.search(None, 'FROM', to[j])
    mail_ids = data[0]
    id_list = mail_ids.split()
    for num in data[0].split():
        typ, data = mail.fetch(num, '(RFC822)' )
        raw_email = data[0][1]
        # converts byte literal to string removing b''
        raw_email_string = raw_email.decode('utf-8')
        email_message = email.message_from_string(raw_email_string)
    if(len(data)>1):
        for response_part in data:
            if isinstance(response_part, tuple):
                msg = email.message_from_string(response_part[1].decode('utf-8'))
                email_subject = msg['subject']
                email_from = msg['from']
                print ('From : ' + email_from + '\n')
                print ('Subject : ' + email_subject + '\n')
                while msg.is_multipart():
                    msg = msg.get_payload(0)
                    content = msg.get_payload(decode=True)
                print(content)
                wbkName = 'dummy.xlsx'
                wbk = openpyxl.load_workbook(wbkName)
                for wks in wbk.worksheets:
                    for myRow in range(1, 100):
                        if(wks.cell(row=myRow, column=6).value == to[j]):
                            wks.cell(row=myRow, column=7).value = content
                            wbk.save(wbkName)
wbk.close

--------------------------------------------------------------------

#def check():
    #wb = Book.caller()
    for j in range(len(to)):
        type, data = mail.search(None, 'FROM', to[j])
        mail_ids = data[0]
        id_list = mail_ids.split()
        for num in data[0].split():
            typ, data = mail.fetch(num, '(RFC822)' )
            raw_email = data[0][1]
            # converts byte literal to string removing b''
            raw_email_string = raw_email.decode('utf-8')
            email_message = email.message_from_string(raw_email_string)
        if(len(data)>1):
            for response_part in data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_string(response_part[1].decode('utf-8'))
                    email_subject = msg['subject']
                    email_from = msg['from']
                    print ('From : ' + email_from + '\n')
                    print ('Subject : ' + email_subject + '\n')
                    while msg.is_multipart():
                        msg = msg.get_payload(0)
                        content = msg.get_payload(decode=True)
                    print(content)
                    wbkName = 'dummy.xlsm'
                    wbk = openpyxl.load_workbook(wbkName)
                    for wks in wbk.worksheets:
                        for myRow in range(1, 100):
                            if(wks.cell(row=myRow, column=6).value == to[j]):
                                wks.cell(row=myRow, column=7).value = content
                                wbk.save(wbkName)
    wbk.close
